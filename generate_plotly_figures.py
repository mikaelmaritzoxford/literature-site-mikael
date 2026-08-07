from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
import plotly.graph_objects as go


NUMERIC_RE = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?")


def first_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    match = NUMERIC_RE.search(text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _col_lookup(headers: list[str]) -> dict[str, int]:
    return {h: i + 1 for i, h in enumerate(headers)}


def _parse_resistivity(row: dict[str, Any]) -> float | None:
    # Prefer the explicit resistivity column.
    value = first_number(row.get("Resistivity (ohm·cm)"))
    if value is not None:
        return value

    # Fall back to the generic property value if it appears to mention resistivity.
    prop_value = _cell_text(row.get("Property value"))
    prop_unit = _cell_text(row.get("Property unit"))
    text = f"{prop_value} {prop_unit}".lower()
    if "ohm" in text or "ω" in text:
        return first_number(prop_value)
    return None


def _parse_mobility(row: dict[str, Any]) -> float | None:
    value = first_number(row.get("Mobility (cm^2/Vs)"))
    if value is not None:
        return value
    prop_value = _cell_text(row.get("Property value"))
    prop_unit = _cell_text(row.get("Property unit"))
    text = f"{prop_value} {prop_unit}".lower()
    if "cm^2/vs" in text or "cm2/vs" in text:
        return first_number(prop_value)
    return None


def _parse_dopant(row: dict[str, Any]) -> float | None:
    value = first_number(row.get("Dopant (at.%)"))
    if value is not None:
        return value

    # Fallback for rows that mention at.% in the property notes.
    comments = _cell_text(row.get("Comments"))
    prop_value = _cell_text(row.get("Property value"))
    for text in (comments, prop_value):
        m = re.search(r"(\d+(?:\.\d+)?)\s*at\.?%", text, flags=re.I)
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                pass
    return None


def _parse_carrier_concentration(row: dict[str, Any]) -> float | None:
    value = first_number(row.get("Carrier concentration (cm^-3)"))
    if value is not None:
        return value
    prop_value = _cell_text(row.get("Property value"))
    text = prop_value.lower().replace("×", "x")
    m = re.search(r"(\d+(?:\.\d+)?)\s*[x×]\s*10\s*\^?\s*([+-]?\d+)", text)
    if m:
        mantissa = float(m.group(1))
        exponent = int(m.group(2))
        return mantissa * (10 ** exponent)
    return None


def generate_dopant_mobility_plot(workbook_path: Path, out_html: Path) -> tuple[Path, int]:
    """Create a Plotly scatter/line plot from the reported data sheet."""
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Reported Data"]

    headers = [_cell_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    col = _col_lookup(headers)

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        values = {h: ws.cell(r, c).value for h, c in col.items()}
        if any(v is not None and str(v).strip() != "" for v in values.values()):
            rows.append(values)

    points = []
    for row in rows:
        dopant = _parse_dopant(row)
        mobility = _parse_mobility(row)
        resistivity = _parse_resistivity(row)
        if dopant is None or mobility is None:
            continue
        points.append({
            "dopant": dopant,
            "mobility": mobility,
            "resistivity": resistivity,
            "paper_id": _cell_text(row.get("Paper ID")),
            "obs_id": _cell_text(row.get("Observation ID")),
            "sample": _cell_text(row.get("Sample / condition")),
            "technique": _cell_text(row.get("Fabrication technique")),
            "temp": first_number(row.get("Deposition temp (C)")),
            "post_temp": first_number(row.get("Post-deposition temp (C)")),
            "carrier": _parse_carrier_concentration(row),
            "sheet_res": first_number(row.get("Sheet resistance (ohm/sq)")),
            "wavelength": first_number(row.get("Wavelength (nm)")),
        })

    out_html.parent.mkdir(parents=True, exist_ok=True)

    if not points:
        fig = go.Figure()
        fig.add_annotation(
            text="No rows currently contain both dopant concentration and mobility.",
            xref="paper",
            yref="paper",
            x=0.5,
            y=0.5,
            showarrow=False,
            font=dict(size=16),
        )
        fig.update_layout(
            template="plotly_white",
            title="IZrO mobility vs dopant concentration",
            xaxis_title="Dopant concentration (at.%)",
            yaxis_title="Mobility (cm²/Vs)",
            height=600,
        )
        fig.write_html(str(out_html), include_plotlyjs="cdn", full_html=True)
        return out_html, 0

    points = sorted(points, key=lambda x: x["dopant"])

    x = [p["dopant"] for p in points]
    y = [p["mobility"] for p in points]
    resistivities = [p["resistivity"] for p in points]

    # For the color scale, use log10 resistivity where available; leave missing values blank.
    color_vals = [math.log10(v) if isinstance(v, (int, float)) and v > 0 else None for v in resistivities]

    customdata = []
    for p in points:
        customdata.append([
            p["paper_id"],
            p["obs_id"],
            p["sample"],
            p["technique"],
            p["resistivity"],
            p["sheet_res"],
            p["carrier"],
            p["temp"],
            p["post_temp"],
            p["wavelength"],
        ])

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=x,
            y=y,
            mode="lines+markers",
            line=dict(width=2),
            marker=dict(
                size=12,
                color=color_vals,
                colorscale="Viridis",
                showscale=True,
                colorbar=dict(title="log10 resistivity\n(Ω·cm)"),
                line=dict(width=1, color="black"),
            ),
            customdata=customdata,
            hovertemplate=(
                "Paper: %{customdata[0]}<br>"
                "Observation: %{customdata[1]}<br>"
                "Sample: %{customdata[2]}<br>"
                "Technique: %{customdata[3]}<br>"
                "Dopant: %{x:.2f} at.%<br>"
                "Mobility: %{y:.2f} cm²/Vs<extra></extra>"
            ),
            name="Reported data",
        )
    )

    fig.update_layout(
        template="plotly_white",
        title="IZrO mobility vs dopant concentration",
        xaxis_title="Dopant concentration (at.%)",
        yaxis_title="Mobility (cm²/Vs)",
        height=650,
        margin=dict(l=60, r=40, t=70, b=60),
    )

    fig.write_html(str(out_html), include_plotlyjs="cdn", full_html=True)
    return out_html, len(points)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate Plotly figures from the literature workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("out_html", type=Path)
    args = parser.parse_args()

    path, n = generate_dopant_mobility_plot(args.workbook, args.out_html)
    print(f"Wrote {path} with {n} point(s)")
