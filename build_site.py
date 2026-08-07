from __future__ import annotations

import argparse
import csv
import importlib
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


FIGURE_RE = re.compile(r"figure\s*(\d+)", re.I)


def cell_str(v):
    return "" if v is None else str(v).strip()


def nonempty(v):
    return v is not None and str(v).strip() != ""


def md_escape(s):
    if s is None:
        return ""
    s = str(s).replace("\r\n", "\n").replace("\r", "\n")
    return s.replace("|", r"\|")


def parse_main(ws):
    """
    Main sheet parser.

    Rules:
    - A1 = site title
    - A row with text in A and empty B = level-1 section
    - A row with text in A and text in B = level-2 subsection, B is first paragraph
    - Rows with text only in B append to the active paragraph block
    - Figure markers can appear in C and are attached to the active paragraph block
    """
    title = ws["A1"].value or "Literature Website"
    sections = []
    current_section = None
    current_sub = None
    current_block = None

    def flush_block():
        nonlocal current_block, current_sub
        if current_block is not None and current_sub is not None:
            current_sub.setdefault("blocks", []).append(current_block)
        current_block = None

    def flush_sub():
        nonlocal current_sub, current_section
        flush_block()
        if current_sub is not None and current_section is not None:
            current_section.setdefault("subsections", []).append(current_sub)
        current_sub = None

    def flush_section():
        nonlocal current_section
        flush_sub()
        if current_section is not None:
            sections.append(current_section)
        current_section = None

    for r in range(2, ws.max_row + 1):
        a = cell_str(ws[f"A{r}"].value)
        b = cell_str(ws[f"B{r}"].value)
        c = cell_str(ws[f"C{r}"].value)

        if not a and not b and not c:
            continue

        if a and not b:
            flush_section()
            current_section = {"heading": a, "intro": [], "subsections": []}
            continue

        if a and b:
            if current_section is None:
                current_section = {"heading": "Untitled", "intro": [], "subsections": []}
            flush_sub()
            current_sub = {"heading": a, "blocks": []}
            current_block = {"lines": [b], "figures": _parse_figures(c)}
            continue

        if not a and b:
            if current_sub is None:
                if current_section is None:
                    current_section = {"heading": "Untitled", "intro": [], "subsections": []}
                current_section.setdefault("intro", []).append({"lines": [b], "figures": _parse_figures(c)})
            else:
                if current_block is None:
                    current_block = {"lines": [b], "figures": _parse_figures(c)}
                else:
                    current_block["lines"].append(b)
                    current_block["figures"].extend(_parse_figures(c))
            continue

        if c:
            if current_block is not None:
                current_block["figures"].extend(_parse_figures(c))

    flush_section()
    return title, sections


def _parse_figures(text: str):
    if not text:
        return []
    parts = re.split(r"[,\n;]+", text)
    out = []
    for part in parts:
        m = FIGURE_RE.search(part.strip())
        if m:
            out.append(int(m.group(1)))
    return out


def _figure_block(relpath: str, caption: str, title: str) -> str:
    return (
        f"**{caption}**\n\n"
        f'<div class="plotly-figure">\n'
        f'  <iframe src="{relpath}" title="{title}" style="width: 100%; height: 720px; border: 0;" loading="lazy"></iframe>\n'
        f"</div>"
    )


def write_index(title, sections, out_path: Path, figure_relpaths: dict[int, str], figure_titles: dict[int, str]):
    lines = [f"# {title}", ""]
    for sec in sections:
        lines.append(f"## {sec['heading']}")
        lines.append("")
        for block in sec.get("intro", []):
            text = " ".join(block.get("lines", []))
            lines.append(md_escape(text))
            lines.append("")
            for fig_num in block.get("figures", []):
                rel = figure_relpaths.get(fig_num)
                if rel:
                    caption = figure_titles.get(fig_num, f"Figure {fig_num}")
                    lines.append(_figure_block(rel, caption, f"Figure {fig_num}"))
                    lines.append("")
        for sub in sec.get("subsections", []):
            lines.append(f"### {sub['heading']}")
            lines.append("")
            for block in sub.get("blocks", []):
                text = " ".join(block.get("lines", []))
                lines.append(md_escape(text))
                lines.append("")
                for fig_num in block.get("figures", []):
                    rel = figure_relpaths.get(fig_num)
                    if rel:
                        caption = figure_titles.get(fig_num, f"Figure {fig_num}")
                        lines.append(_figure_block(rel, caption, f"Figure {fig_num}"))
                        lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")


def _sheet_headers_and_rows(ws):
    max_col = ws.max_column
    max_row = ws.max_row

    header_row = None
    headers = []
    rows = []
    for r in range(1, max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if any(nonempty(v) for v in vals):
            header_row = r
            headers = [cell_str(v) if nonempty(v) else f"Column {i}" for i, v in enumerate(vals, start=1)]
            break

    if header_row is None:
        return [], []

    for r in range(header_row + 1, max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if not any(nonempty(v) for v in vals):
            continue
        row = [cell_str(v) for v in vals[:len(headers)]]
        if len(row) < len(headers):
            row.extend([""] * (len(headers) - len(row)))
        rows.append(row)

    return headers, rows


def write_sheet_csv(ws, csv_path: Path):
    headers, rows = _sheet_headers_and_rows(ws)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if headers:
            writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


def write_grid_page(
    title: str,
    csv_relpath: str,
    out_path: Path,
    notes: str | None = None,
    column_picker: bool = False,
):
    lines = [f"# {title}", ""]
    if notes:
        lines.append(notes)
        lines.append("")
    picker_attr = ' data-column-picker="true"' if column_picker else ""
    lines.append(
        f'<div class="csv-grid" data-csv="{csv_relpath}" '
        f'data-title="{title}"{picker_attr}></div>'
    )
    lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")


def write_text_page(ws, title: str, out_path: Path):
    headers, rows = _sheet_headers_and_rows(ws)
    if not headers:
        out_path.write_text(f"# {title}\n\n_No data found._\n", encoding="utf-8")
        return

    lines = [f"# {title}", ""]
    lines.append("| " + " | ".join(md_escape(h) for h in headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        lines.append("| " + " | ".join(md_escape(x).replace("\n", "<br>") for x in row) + " |")
    lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")


def collect_figures(ws):
    figs = set()
    for r in range(1, ws.max_row + 1):
        c = cell_str(ws[f"C{r}"].value)
        figs.update(_parse_figures(c))
    return figs


def generate_figures(workbook_path: Path, docs_dir: Path, figures: set[int]) -> tuple[dict[int, str], dict[int, str]]:
    relpaths = {}
    titles = {}
    plots_dir = docs_dir / "assets" / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    for n in sorted(figures):
        module_name = f"figures.figure{n}"
        try:
            module = importlib.import_module(module_name)
        except Exception as exc:
            raise RuntimeError(f"Could not import {module_name}: {exc}") from exc

        generator = getattr(module, "generate_figure", None)
        if generator is None:
            raise AttributeError(f"{module_name} must define generate_figure(workbook_path, out_html)")
        out_html = plots_dir / f"figure{n}.html"
        generator(workbook_path, out_html)
        relpaths[n] = f"assets/plots/figure{n}.html"
        titles[n] = getattr(module, "FIGURE_TITLE", f"Figure {n}")

    return relpaths, titles


def main():
    parser = argparse.ArgumentParser(description="Generate MkDocs pages and CSV grids from an Excel workbook.")
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx workbook")
    parser.add_argument("--project-dir", type=Path, default=Path(__file__).resolve().parent, help="Project root")
    parser.add_argument("--build", action="store_true", help="Run mkdocs build after generating pages")
    args = parser.parse_args()

    project_dir = args.project_dir.resolve()
    docs_dir = project_dir / "docs"
    docs_dir.mkdir(exist_ok=True)

    wb = load_workbook(args.workbook, data_only=True)
    main_ws = wb["Main"]

    # Generate figures first so the index can reference them.
    main_title, main_sections = parse_main(main_ws)
    figure_relpaths, figure_titles = generate_figures(args.workbook, docs_dir, collect_figures(main_ws))
    write_index(main_title, main_sections, docs_dir / "index.md", figure_relpaths, figure_titles)

    # Export sheet data to CSV and create grid pages.
    grid_sheets = [
        ("Papers", "papers", False),
        ("Reported Data", "reported-data", True),
    ]

    for sheet_name, slug, column_picker in grid_sheets:
        csv_path = docs_dir / "data" / f"{slug}.csv"
        write_sheet_csv(wb[sheet_name], csv_path)
        write_grid_page(
            sheet_name,
            # Pages publish at /<slug>/index.html; CSV files publish at /data/.
            f"../data/{slug}.csv",
            docs_dir / f"{slug}.md",
            notes="Interactive table. Use the search box, sort headers, resize columns, and scroll horizontally as needed.",
            column_picker=column_picker,
        )

    if args.build:
        subprocess.run([sys.executable, "-m", "mkdocs", "build"], cwd=project_dir, check=True)


if __name__ == "__main__":
    main()
